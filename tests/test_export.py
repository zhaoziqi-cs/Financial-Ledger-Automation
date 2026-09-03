from io import BytesIO

import openpyxl

from ledger_etl import export, pipeline


def test_export_bytes_openable(env):
    env.seed()
    env.write_flow()
    pipeline.run_flow(env.settings, env.bank_file)  # 主表 6 行 + 1 条待处理未识别

    lb = export.ledger_to_xlsx_bytes(env.settings)
    wb = openpyxl.load_workbook(BytesIO(lb))
    ws = wb.active
    assert ws.max_row - 1 == 6  # 去掉表头

    ub = export.unmatched_to_xlsx_bytes(env.settings, resolved=False)
    wbu = openpyxl.load_workbook(BytesIO(ub))
    assert wbu.active.max_row - 1 == 1  # 1 条待处理


if __name__ == "__main__":
    import sys
    import pytest

    # 方便单独试验：python tests/test_xxx.py [-k foo]
    sys.exit(pytest.main([__file__, *sys.argv[1:]]))
