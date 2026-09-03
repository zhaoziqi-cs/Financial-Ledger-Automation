"""导出 xlsx bytes：主表 / 未识别清单。"""
from __future__ import annotations

from io import BytesIO

import pandas as pd

from . import db
from .config import Settings

MS_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="台账")
    return bio.getvalue()


def ledger_to_xlsx_bytes(settings: Settings) -> bytes:
    conn = db.connect(settings.db_path)
    try:
        df = db.read_ledger(conn)
    finally:
        conn.close()
    return df_to_xlsx_bytes(df)


def unmatched_to_xlsx_bytes(settings: Settings, resolved: bool | None = False) -> bytes:
    """导出异常行。resolved=None 导出全部；False=待处理(open+conflict)；True=已解决。"""
    conn = db.connect(settings.db_path)
    try:
        df = db.read_unmatched(conn)
    finally:
        conn.close()
    if resolved is True:
        df = df[df["status"] == "resolved"]
    elif resolved is False:
        df = df[df["status"] != "resolved"]
    return df_to_xlsx_bytes(df)


if __name__ == "__main__":
    # 运行：python -m ledger_etl.export
    from io import BytesIO

    import openpyxl

    print("## export：把 DataFrame 导出成 .xlsx（内存 bytes）")
    demo = pd.DataFrame([
        {"date": "2026-04-29", "project": "百家坊未来社区", "summary": "支付工程款", "income": 0.0, "expense": 100.0, "balance": -100.0},
        {"date": "2026-04-30", "project": "丽水人民医院", "summary": "支付工程款", "income": 0.0, "expense": 50.0, "balance": -150.0},
    ])
    data = df_to_xlsx_bytes(demo)
    wb = openpyxl.load_workbook(BytesIO(data))
    print("  字节数 =", len(data), "| 工作表 =", wb.sheetnames, "| 行数(含表头) =", wb.active.max_row)
    print("真实导出请用：ledger export-ledger / ledger export-audit")
